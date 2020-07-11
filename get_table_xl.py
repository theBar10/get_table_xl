import openpyxl
import pandas as pd


wb = openpyxl.load_workbook(r'C:\Users\tbarten\Desktop\Projects\Cost_Productivity\2020 EFCO Productivity.xlsm', data_only = True)
"""
To use with a different file, change the file path in the above assignment for wb. Use r' to read the file.
Using data_only allows for the script to pull the calculated value in Excel, without it it will show the underlining formula
Commented out print(wb.get_sheet_names()) to only use when I need to check the sheet names for the next step
"""
#print(wb.get_sheet_names()) 

#selecting the tab from the excel file that I want to work in; tab name is a string; keep it in ('')
sheet = wb.get_sheet_by_name('Actual Drivers')
#the tables in the sheets do not align as well as I thought they would, should make a copy of this script and change the sheet and values to match

"""myDict is set to a blank dictionary in prep for the series of list that the next section will add.

Using the below outer for and inner for loops you can work through an array of cells in Excel to get their values.
The outer for loop (i) sets the number of rows to get the values from.
The inner for loop (n) sets the number of columns to get the values from.
With each for loop the first integer in the range is the cell number to start with, it must start with 1 or higher (zero doesn't equal one).
The second integer in the range for the loops is the number of rows or columns that are to be looped through."""

atable_dict = \
{'Reissues':{'first': 6, 'last':11, 'uom': 'dollars', 'group_type': 'type'}, \
'Pounds':{'first': 16, 'last':22, 'uom': 'pounds', 'group_type': 'department'}, \
'Window_Units':{'first': 61, 'last':70, 'uom': 'units', 'group_type': 'department'}, \
'Warehouse_Pieces':{'first': 95, 'last':96, 'uom': 'sticks', 'group_type': 'department'}, \
'Downtown_Units':{'first': 75, 'last':83, 'uom': 'units', 'group_type': 'department'}, \
'Glass_SQFT':{'first': 83, 'last':88, 'uom': 'SQFT', 'group_type': 'department'}, \
'Grid_Units':{'first': 88, 'last':89, 'uom': 'units', 'group_type': 'department'}, \
'Finish_SQFT':{'first': 97, 'last':99, 'uom': 'SQFT', 'group_type': 'department'}, \
'Thermal_Pieces':{'first': 99, 'last':101, 'uom': 'sticks', 'group_type': 'department'}, \
'Window_Units_Straight':{'first': 139, 'last':145, 'uom': 'units', 'group_type': 'department'}, \
'Window_Hours':{'first': 149, 'last':158, 'uom': 'hours', 'group_type': 'department'}, \
'Downtown_Hours':{'first': 163, 'last':171, 'uom': 'hours', 'group_type': 'department'}, \
'Glass_Hours':{'first': 171, 'last':176, 'uom': 'hours', 'group_type': 'department'}, \
'Grid_Hours':{'first': 176, 'last':177, 'uom': 'hours', 'group_type': 'department'}, \
'Extrusion_Hours':{'first': 182, 'last':183, 'uom': 'hours', 'group_type': 'department'}, \
'Warehouse_Hours':{'first': 183, 'last':185, 'uom': 'hours', 'group_type': 'department'}, \
'Finish_Hours':{'first': 185, 'last':187, 'uom': 'hours', 'group_type': 'department'}, \
'Thermal_Hours':{'first': 187, 'last':189, 'uom': 'hours', 'group_type': 'department'}, \
'Indirect_Hours':{'first': 195, 'last':204, 'uom': 'hours', 'group_type': 'department'}, \
'Window_Hours_OT':{'first': 228, 'last':237, 'uom': 'hours_ot', 'group_type': 'department'}, \
'Downtown_Hours_OT':{'first': 242, 'last':250, 'uom': 'hours_ot', 'group_type': 'department'}, \
'Glass_Hours_OT':{'first': 250, 'last':255, 'uom': 'hours_ot', 'group_type': 'department'}, \
'Grid_Hours_OT':{'first': 255, 'last':256, 'uom': 'hours_ot', 'group_type': 'department'}, \
'Extrusion_Hours_OT':{'first': 261, 'last':262, 'uom': 'hours_ot', 'group_type': 'department'}, \
'Warehouse_Hours_OT':{'first': 262, 'last':264, 'uom': 'hours_ot', 'group_type': 'department'}, \
'Finish_Hours_OT':{'first': 264, 'last':266, 'uom': 'hours_ot', 'group_type': 'department'}, \
'Thermal_Hours_OT':{'first': 266, 'last':268, 'uom': 'hours_ot', 'group_type': 'department'}, \
'Indirect_Hours_OT':{'first': 274, 'last':283, 'uom': 'hours_ot', 'group_type': 'department'}, \
'COQ':{'first': 307, 'last':310, 'uom': 'dollars', 'group_type': 'type'}}


for table in atable_dict:
    myDict = {}
    for i in range (5, 6): #for the fixed row of dates
            for n in range (2, 74): 
                data = ([])
                data.append(sheet.cell(row = i, column = n).value)
                myDict.setdefault(n, []).append(data)

        
    for i in range (atable_dict[table]['first'], atable_dict[table]['last']):
        for n in range (2, 74): 
            data = ([])
            data.append(sheet.cell(row = i, column = n).value)
            myDict.setdefault(n, []).append(data)

    delete = (3,4,5,6,7)

    for d in delete:
        del myDict[d]
    
    """Pandas (pd) are easier to work with when reshaping data for analysis.
    I wanted to use the pandas dataframe to wrangle the data."""    
    df = pd.DataFrame.from_dict(myDict)
    
    """"Since I created the dataframe out of list of list I was stuck with square brackets
    This for loop gets the string of list, effectively removing the square brackets from the dataframe."""
    for r in df:
        df[r] = df[r].str.get(0)
    
 
    """I am sure there is a much cleaner way to reshape a long data set that has a column and row that need to be indexed.
    Either way, I modified the data into a long form with stack, then I unstacked what was the first row, then I used those
    two columns to 'melt' the data into the shape that I was trying to work with."""     
    df=df.set_index([2]).stack().unstack([0]).reset_index()
    df=df.rename(columns={None:'Dates'})
    df=df.melt(id_vars = ['index', 'Dates'])
    
    """Added in some dynamic column name changes. 
    This helps drill with the two common variables of this dataset."""
    group = atable_dict[table]['group_type']
    uom = atable_dict[table]['uom']
    
    df=df.rename(columns={2:group})
    df=df.rename(columns={'value':uom})
    
        
    """The data set had this reoccurring month to date (MTD) 
    and year to date (YTD) calculation in it that was not needed. 
    The doesn't equal removes or filters out that calculation."""
    df=df[df.Dates != 'MTD']
    df=df[df.Dates != 'YTD']
    
    df.fillna(0) #making null values zero
    """Adding in a new column with a repeating value to separate budget data from actual data"""    
    df=df.assign(productivity='Actual')
    #print(df.info())
    
    """save the reshaped dataframe and changed the name to dynamically align with dict key"""
    df.to_csv(r'C:\Users\tbarten\Desktop\Projects\Cost_Productivity\PyRun\Actual 2020 {0}.csv'.format(table))

"""Just repeating the entire process but with a different sheet name. 
This could have been done with another loop but the tables are in slightly different locations so I thought it
would be quicker to copy the first top of the code"""
#selecting the tab from the excel file that I want to work in; tab name is a string; keep it in ('')
sheet_b = wb.get_sheet_by_name('Budget Drivers')
#the tables in the sheets do not align as well as I thought they would, should make a copy of this script and change the sheet and values to match

"""myDict is set to a blank dictionary in prep for the series of list that the next section will add.

Using the below outer for and inner for loops you can work through an array of cells in Excel to get their values.
The outer for loop (i) sets the number of rows to get the values from.
The inner for loop (n) sets the number of columns to get the values from.
With each for loop the first integer in the range is the cell number to start with, it must start with 1 or higher (zero doesn't equal one).
The second integer in the range for the loops is the number of rows or columns that are to be looped through."""

btable_dict = \
{'Reissues':{'first': 7, 'last':11, 'uom': 'dollars', 'group_type': 'type'}, \
'Pounds':{'first': 16, 'last':18, 'uom': 'pounds', 'group_type': 'type'}, \
'Window_Units':{'first': 19, 'last':28, 'uom': 'units', 'group_type': 'department'}, \
'Downtown_Units':{'first': 33, 'last':41, 'uom': 'units', 'group_type': 'department'}, \
'Glass_SQFT':{'first': 41, 'last':46, 'uom': 'SQFT', 'group_type': 'department'}, \
'Grid_Units':{'first': 46, 'last':47, 'uom': 'units', 'group_type': 'department'}, \
'Warehouse_Pieces':{'first': 53, 'last':55, 'uom': 'sticks', 'group_type': 'department'}, \
'Finish_SQFT':{'first': 55, 'last':57, 'uom': 'SQFT', 'group_type': 'department'}, \
'Thermal_Pieces':{'first': 57, 'last':59, 'uom': 'sticks', 'group_type': 'department'}, \
'Window_Hours':{'first': 98, 'last':107, 'uom': 'hours', 'group_type': 'department'}, \
'Downtown_Hours':{'first': 112, 'last':120, 'uom': 'hours', 'group_type': 'department'}, \
'Glass_Hours':{'first': 120, 'last':125, 'uom': 'hours', 'group_type': 'department'}, \
'Grid_Hours':{'first': 125, 'last':126, 'uom': 'hours', 'group_type': 'department'}, \
'Extrusion_Hours':{'first': 131, 'last':132, 'uom': 'hours', 'group_type': 'department'}, \
'Warehouse_Hours':{'first': 132, 'last':134, 'uom': 'hours', 'group_type': 'department'}, \
'Finish_Hours':{'first': 134, 'last':136, 'uom': 'hours', 'group_type': 'department'}, \
'Thermal_Hours':{'first': 136, 'last':138, 'uom': 'hours', 'group_type': 'department'}, \
'Indirect_Hours':{'first': 144, 'last':153, 'uom': 'hours', 'group_type': 'department'}, \
'Window_Hours_OT':{'first': 177, 'last':186, 'uom': 'hours_ot', 'group_type': 'department'}, \
'Downtown_Hours_OT':{'first': 191, 'last':199, 'uom': 'hours_ot', 'group_type': 'department'}, \
'Glass_Hours_OT':{'first': 199, 'last':204, 'uom': 'hours_ot', 'group_type': 'department'}, \
'Grid_Hours_OT':{'first': 204, 'last':205, 'uom': 'hours_ot', 'group_type': 'department'}, \
'Extrusion_Hours_OT':{'first': 210, 'last':211, 'uom': 'hours_ot', 'group_type': 'department'}, \
'Warehouse_Hours_OT':{'first': 211, 'last':213, 'uom': 'hours_ot', 'group_type': 'department'}, \
'Finish_Hours_OT':{'first': 213, 'last':215, 'uom': 'hours_ot', 'group_type': 'department'}, \
'Thermal_Hours_OT':{'first': 215, 'last':217, 'uom': 'hours_ot', 'group_type': 'department'}, \
'Indirect_Hours_OT':{'first': 223, 'last':232, 'uom': 'hours_ot', 'group_type': 'department'}, \
'COQ':{'first': 256, 'last':259, 'uom': 'dollars', 'group_type': 'type'}}


for table in btable_dict:
    myDict = {}
    for i in range (5, 6): #for the fixed row of dates
            for n in range (2, 74): 
                data = ([])
                data.append(sheet_b.cell(row = i, column = n).value)
                myDict.setdefault(n, []).append(data)

        
    for i in range (btable_dict[table]['first'], btable_dict[table]['last']):
        for n in range (2, 74): 
            data = ([])
            data.append(sheet_b.cell(row = i, column = n).value)
            myDict.setdefault(n, []).append(data)

    delete = (3,4,5,6,7)

    for d in delete:
        del myDict[d]
    
    """Pandas (pd) are easier to work with when reshaping data for analysis.
    I wanted to use the pandas dataframe to wrangle the data."""    
    df = pd.DataFrame.from_dict(myDict)
    
    """"Since I created the dataframe out of list of list I was stuck with square brackets
    This for loop gets the string of list, effectively removing the square brackets from the dataframe."""
    for r in df:
        df[r] = df[r].str.get(0)
    
    #print(df.head())
    """I am sure there is a much cleaner way to reshape a long data set that has a column and row that need to be indexed.
    Either way, I modified the data into a long form with stack, then I unstacked what was the first row, then I used those
    two columns to 'melt' the data into the shape that I was trying to work with."""     
    df=df.set_index([2]).stack().unstack([0]).reset_index()
    df=df.rename(columns={None:'Dates'})
    df=df.melt(id_vars = ['index', 'Dates'])
    
    """Added in some dynamic column name changes. 
    This helps drill with the two common variables of this dataset."""
    group = btable_dict[table]['group_type']
    uom = btable_dict[table]['uom']
    
    df=df.rename(columns={2:group})
    df=df.rename(columns={'value':uom})
    
    """Adding in a new column with a repeating value to separate budget data from actual data"""
    df=df.assign(productivity='Budget')
    
    """The data set had this reoccurring month to date (MTD) 
    and year to date (YTD) calculation in it that was not needed. 
    The doesn't equal removes or filters out that calculation."""
    df=df[df.Dates != 'MTD']
    df=df[df.Dates != 'YTD']
    
    df.fillna(0) #making null values zero
    
    #print(df.info())
    
    """save the reshaped dataframe and changed the name to dynamically align with dict key"""
    df.to_csv(r'C:\Users\tbarten\Desktop\Projects\Cost_Productivity\PyRun\Budget 2020 {0}.csv'.format(table))
